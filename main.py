# main.py
import asyncio
import logging
import json
from datetime import datetime
from pathlib import Path

import websockets
import aiohttp

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import config

from src.scanner import process_kline, prefill_history
from src.research import aggregate_signals
from src.research import fetch_cryptopanic, fetch_fear_greed, fetch_rss, analyze_with_claude
from src.risk import load_state, save_state, check_risk
from src.executor import execute_signal
from src.position_monitor import monitor_positions, load_positions
from src.telegram_bot import notify_signal, notify_startup, notify_cryptopanic_disabled, notify_position_closed, notify_shutdown

Path("logs").mkdir(exist_ok=True)

_fmt     = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
_sh      = logging.StreamHandler()
_sh.setFormatter(_fmt)
_fh      = logging.FileHandler("logs/main.log", encoding="utf-8")
_fh.setFormatter(_fmt)

root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)
root_logger.addHandler(_sh)
root_logger.addHandler(_fh)

log = logging.getLogger(__name__)

current_trade_signal = {"signal": "neutral", "confidence": 0.0, "fear_greed": {"value": 50, "label": "Neutral"}}
last_anomaly_flag    = False
last_entry_time      = 0.0
last_telegram_time   = 0.0
ENTRY_COOLDOWN       = 1200
TELEGRAM_COOLDOWN    = 1800

# Shared state for monitor_task - updated from scanner_task
current_price_ref = {"price": 0.0}

TRADES_FILE = Path("logs/trades.xlsx")


def log_trade(
    signal: str,
    price: float,
    position_size: float,
    stop_loss: float,
    confidence: float,
    z_score: float,
    order_id: str
):
    import openpyxl

    xlsx_file = Path("logs/trades.xlsx")

    if xlsx_file.exists():
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "trades"
        ws.append([
            "timestamp", "signal", "price_entry",
            "position_size", "stop_loss", "confidence",
            "z_score", "order_id",
            "price_exit", "pnl", "pnl_pct", "close_reason", "closed_at"
        ])

    ws.append([
        datetime.utcnow().isoformat(),
        signal,
        round(price, 2),
        round(position_size, 2),
        round(stop_loss, 2),
        round(confidence, 2),
        round(z_score, 2),
        order_id
    ])

    wb.save(xlsx_file)
    log.info(f"The deal is saved | {signal.upper()} @ ${price:,.2f}")


def log_trade_close(
    order_id: str,
    signal: str,
    price_entry: float,
    price_exit: float,
    position_size: float,
    pnl: float,
    pnl_pct: float,
    reason: str
):
    import openpyxl

    xlsx_file = Path("logs/trades.xlsx")

    if xlsx_file.exists():
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "trades"
        ws.append([
            "timestamp", "signal", "price_entry",
            "position_size", "stop_loss", "confidence",
            "z_score", "order_id",
            "price_exit", "pnl", "pnl_pct", "close_reason", "closed_at"
        ])

    # Find row by order_id and update it
    updated = False
    for row in ws.iter_rows(min_row=2):
        if row[7].value == order_id:
            row[8].value = round(price_exit, 2)
            row[9].value = round(pnl, 4)
            row[10].value = round(pnl_pct, 2)
            row[11].value = reason
            row[12].value = datetime.utcnow().isoformat()
            updated = True
            break

    if not updated:
        # Row not found — append as new row
        ws.append([
            datetime.utcnow().isoformat(),
            signal,
            round(price_entry, 2),
            round(position_size, 2),
            "", "", "",
            order_id,
            round(price_exit, 2),
            round(pnl, 4),
            round(pnl_pct, 2),
            reason,
            datetime.utcnow().isoformat()
        ])

    wb.save(xlsx_file)
    log.info(f"Close saved | {order_id} | PnL: {pnl_pct:+.2f}%")


async def research_task():
    global current_trade_signal

    async with aiohttp.ClientSession() as session:
        while True:
            try:
                log.info("Research: getting news...")
                news1 = await fetch_cryptopanic(session)
                news2 = await fetch_rss(session)
                fg    = await fetch_fear_greed(session)

                all_news = news1 + news2
                signals  = []

                for item in all_news:
                    signal = await analyze_with_claude(
                        session, item["title"], item["content"], item["source"]
                    )
                    if signal:
                        signals.append(signal)

                trade_signal               = aggregate_signals(signals, fg)
                trade_signal["fear_greed"] = fg
                current_trade_signal       = trade_signal

                log.info(
                    f"Research result: {trade_signal['signal'].upper()} | "
                    f"Confidence: {trade_signal['confidence']} | "
                    f"Fear&Greed: {fg['value']} ({fg['label']})"
                )

            except Exception as e:
                log.error(f"Research error: {e}")

            await asyncio.sleep(300)


async def scanner_task():
    global last_anomaly_flag, last_entry_time, last_telegram_time

    from src.scanner import WS_URL
    log.info(f"Scanner: connecting to {WS_URL}")

    while True:
        try:
            # Prefill history before connecting — instant warmup after reconnect
            prefill_history()

            async with websockets.connect(
                WS_URL,
                ping_interval=20,
                ping_timeout=10
            ) as ws:
                log.info("Scanner: WebSocket connected")
                async for message in ws:
                    data = json.loads(message)
                    sig  = await process_kline(data)

                    if sig is None:
                        continue

                    # Update shared price for monitor_task
                    current_price_ref["price"] = sig.price
                    last_anomaly_flag          = sig.anomaly

                    # Entry only on full signal (all filters passed)
                    if sig.signal == "neutral":
                        continue

                    # Research signal must agree with scanner direction
                    research_signal = current_trade_signal["signal"]
                    if research_signal == "neutral" or research_signal != sig.signal:
                        log.info(
                            f"Signal direction mismatch | "
                            f"Scanner: {sig.signal.upper()} | "
                            f"Research: {research_signal.upper()} | Skipping"
                        )
                        continue

                    now_t = asyncio.get_event_loop().time()
                    if now_t - last_entry_time < ENTRY_COOLDOWN:
                        continue

                    last_entry_time = now_t
                    confidence      = current_trade_signal["confidence"]

                    log.info(
                        f"ENTRY SIGNAL | {sig.signal.upper()} | "
                        f"Z: {sig.z_score:+.2f} | "
                        f"RSI: {sig.rsi} | "
                        f"EMA: {sig.trend} | "
                        f"Vol: {'OK' if sig.volume_confirmed else 'LOW'} | "
                        f"Confidence: {confidence}"
                    )

                    state    = load_state()
                    decision = check_risk(
                        state=state,
                        signal=sig.signal,
                        confidence=confidence,
                        current_price=sig.price
                    )

                    if decision.allowed:
                        result = await execute_signal(
                            signal=sig.signal,
                            decision=decision,
                            state=state,
                            price=sig.price,
                            z_score=sig.z_score,
                            confidence=confidence
                        )
                        if result and result.success:
                            log_trade(
                                signal=sig.signal,
                                price=result.price,
                                position_size=result.usdt_value,
                                stop_loss=decision.stop_loss,
                                confidence=confidence,
                                z_score=sig.z_score,
                                order_id=result.order_id
                            )
                    else:
                        log.warning(f"Risk declined: {decision.reason}")

                    # Telegram — no more than once per hour
                    now_tg = asyncio.get_event_loop().time()
                    if now_tg - last_telegram_time >= TELEGRAM_COOLDOWN:
                        last_telegram_time = now_tg
                        await notify_signal(
                            signal=sig.signal,
                            price=sig.price,
                            stop_loss=decision.stop_loss,
                            z_score=sig.z_score,
                            confidence=confidence,
                            fear_greed_value=current_trade_signal["fear_greed"]["value"],
                            fear_greed_label=current_trade_signal["fear_greed"]["label"]
                        )

        except Exception as e:
            log.error(f"Scanner error: {e} | Reconnecting in 5 seconds...")
            await asyncio.sleep(5)


async def monitor_task():
    """Monitors open positions for TP / SL / Reverse signal."""
    await monitor_positions(
        get_price_fn=lambda: current_price_ref["price"],
        get_signal_fn=lambda: current_trade_signal["signal"],
        notify_fn=notify_position_closed,
        on_close_fn=log_trade_close
    )


async def status_task():
    while True:
        await asyncio.sleep(30)
        state  = load_state()
        signal = current_trade_signal

        log.info(
            f"[STATUS] "
            f"Balance: ${state.balance:.2f} | "
            f"Positions: {state.open_positions} | "
            f"Deals: {state.total_trades} | "
            f"Signal: {signal['signal'].upper()} | "
            f"Anomaly: {last_anomaly_flag}"
        )


async def main():
    log.info("=" * 50)
    log.info("BTC Trading Bot active")
    log.info(f"Mode: {config.MODE} | Symbol: {config.SYMBOL}")
    log.info("=" * 50)

    state = load_state()
    state.open_positions = 0
    save_state(state)
    log.info(f"Positions reset | Balance: ${state.balance:.2f}")

    await notify_startup(config.MODE, state.balance)

    await asyncio.gather(
        research_task(),
        scanner_task(),
        monitor_task(),
        status_task()
    )


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        log.info("KeyboardInterrupt received - shutting down")
        state = load_state()
        open_positions = len(load_positions())
        asyncio.run(
            notify_shutdown(
                balance=state.balance,
                open_positions=open_positions,
                total_trades=state.total_trades,
            )
        )
